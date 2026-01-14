"""
Step 2: Canonical Mapping (CanonicalMapper Class)

⚠️⚠️⚠️ STEP2-B DISABLED ⚠️⚠️⚠️

This module is DISABLED as of STEP PIPELINE-V2-BLOCK-STEP2B-01.

REASON: coverage_code must come from SSOT input (Step1 V2), not from string-based assignment.

DO NOT USE THIS MODULE. It will raise an exception if instantiated.

CONSTITUTIONAL VIOLATION:
- coverage_code generation via string matching is FORBIDDEN
- Contaminated path (data/sources/mapping/) is FORBIDDEN
- Use Step1 V2 (pipeline.step1_targeted_v2) instead

⚠️⚠️⚠️ STEP2-B DISABLED ⚠️⚠️⚠️
"""

import csv
import re
from pathlib import Path
from typing import Dict, List, Optional
import openpyxl


class CanonicalMapper:
    """
    담보명 mapping 엑셀 기반 canonical 매핑

    ⚠️⚠️⚠️ DISABLED - DO NOT USE ⚠️⚠️⚠️
    """

    def __init__(self, mapping_excel_path: str):
        raise RuntimeError(
            "STEP2-B DISABLED: coverage_code must come from SSOT input. "
            "String-based assignment is forbidden. "
            "Use Step1 V2 (pipeline.step1_targeted_v2) instead. "
            "[STEP PIPELINE-V2-BLOCK-STEP2B-01]"
        )

        # DEAD CODE BELOW (never executed)
        self.mapping_excel_path = Path(mapping_excel_path)
        self.mapping_dict: Dict[str, Dict] = {}
        self._load_mapping()

    def _normalize(self, text: str) -> str:
        """
        텍스트 정규화 (공백, 특수문자 제거)

        Args:
            text: 원본 텍스트

        Returns:
            str: 정규화된 텍스트 (소문자, 공백/특수문자 제거)
        """
        # 공백 제거
        text = re.sub(r'\s+', '', text)
        # 특수문자 제거 (한글, 영문, 숫자만 유지)
        text = re.sub(r'[^가-힣a-zA-Z0-9]', '', text)
        return text.lower()

    def _load_mapping(self):
        """
        담보명 mapping 엑셀 로드

        엑셀 구조:
        - ins_cd: 보험사 코드
        - 보험사명: 보험사 이름
        - cre_cvr_cd: 신정원 코드 (coverage_code)
        - 신정원코드명: 표준 담보명 (coverage_name_canonical)
        - 담보명(가입설계서): 보험사별 담보명
        """
        if not self.mapping_excel_path.exists():
            raise FileNotFoundError(f"Mapping excel not found: {self.mapping_excel_path}")

        wb = openpyxl.load_workbook(self.mapping_excel_path, data_only=True)
        ws = wb.active

        # 헤더 읽기
        headers = [cell.value for cell in ws[1]]

        # 데이터 로드
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # ins_cd가 없으면 스킵
                continue

            row_data = dict(zip(headers, row))

            # 실제 컬럼명 사용
            coverage_code = str(row_data.get('cre_cvr_cd', '')).strip()
            coverage_name_canonical = str(row_data.get('신정원코드명', '')).strip()
            coverage_name_insurer = str(row_data.get('담보명(가입설계서)', '')).strip()

            if not coverage_code or not coverage_name_canonical:
                continue

            # 1. 신정원코드명으로 exact match
            self.mapping_dict[coverage_name_canonical] = {
                'coverage_code': coverage_code,
                'coverage_name_canonical': coverage_name_canonical,
                'match_type': 'exact'
            }

            # 2. 신정원코드명 normalized match
            normalized_canonical = self._normalize(coverage_name_canonical)
            if normalized_canonical:
                self.mapping_dict[normalized_canonical] = {
                    'coverage_code': coverage_code,
                    'coverage_name_canonical': coverage_name_canonical,
                    'match_type': 'normalized'
                }

            # 3. 보험사별 담보명(가입설계서)으로 exact match
            if coverage_name_insurer:
                self.mapping_dict[coverage_name_insurer] = {
                    'coverage_code': coverage_code,
                    'coverage_name_canonical': coverage_name_canonical,
                    'match_type': 'alias'
                }

                # 4. 보험사별 담보명 normalized match
                normalized_insurer = self._normalize(coverage_name_insurer)
                if normalized_insurer:
                    self.mapping_dict[normalized_insurer] = {
                        'coverage_code': coverage_code,
                        'coverage_name_canonical': coverage_name_canonical,
                        'match_type': 'normalized_alias'
                    }

    def _remove_suffix_patterns(self, text: str) -> Optional[str]:
        """
        Remove trailing suffix patterns (condition/period/limit metadata)

        STEP NEXT-27: Evidence-bound suffix removal for canonical recovery

        Allowed Patterns (evidence-verified):
        - (1년50%), (1년주기): period metadata
        - (최초1회한): occurrence limit
        - (1일-180일): duration range
        - (갱신형_10년): renewal metadata
        - (특정심장질환), (뇌졸중): condition specifier (when base coverage exists)
        - (재진단형): diagnosis type (ARTIFACT case - excluded)

        Args:
            text: Original coverage name with suffix

        Returns:
            Suffix-removed text, or None if no suffix pattern matched
        """
        # Pattern 1: Period/occurrence metadata
        # Examples: (1년50%), (최초1회한), (1년주기)
        pattern_period = r'\((?:\d+년\d+%?|\d+년주기|최초\d+회한)\)$'
        text_stripped = re.sub(pattern_period, '', text).strip()
        if text_stripped != text:
            return text_stripped

        # Pattern 2: Duration range
        # Examples: (1일-180일)
        pattern_duration = r'\(\d+일-\d+일\)$'
        text_stripped = re.sub(pattern_duration, '', text).strip()
        if text_stripped != text:
            return text_stripped

        # Pattern 3: Renewal metadata
        # Examples: (갱신형_10년)
        pattern_renewal = r'\(갱신형_\d+년\)$'
        text_stripped = re.sub(pattern_renewal, '', text).strip()
        if text_stripped != text:
            return text_stripped

        # Pattern 4: Condition specifier (only if base coverage exists in mapping)
        # Examples: (특정심장질환), (뇌졸중)
        # IMPORTANT: Only for cases where Excel has base coverage without condition
        pattern_condition = r'\([^)]+질환\)$|\(뇌졸중\)$'
        text_stripped = re.sub(pattern_condition, '', text).strip()
        if text_stripped != text:
            return text_stripped

        # No pattern matched
        return None

    def map_coverage(self, coverage_name_raw: str) -> Dict:
        """
        담보명을 canonical로 매핑

        Args:
            coverage_name_raw: 원본 담보명

        Returns:
            dict: {
                'coverage_code': str,
                'coverage_name_canonical': str,
                'mapping_status': 'matched' | 'unmatched',
                'match_type': str
            }
        """
        # 1. Exact match
        if coverage_name_raw in self.mapping_dict:
            result = self.mapping_dict[coverage_name_raw].copy()
            result['mapping_status'] = 'matched'
            return result

        # 2. Normalized match
        normalized = self._normalize(coverage_name_raw)
        if normalized in self.mapping_dict:
            result = self.mapping_dict[normalized].copy()
            result['mapping_status'] = 'matched'
            return result

        # 3. Suffix-normalized match (STEP NEXT-27)
        # Only attempt if coverage has suffix pattern
        suffix_removed = self._remove_suffix_patterns(coverage_name_raw)
        if suffix_removed:
            # Try exact match on suffix-removed string
            if suffix_removed in self.mapping_dict:
                result = self.mapping_dict[suffix_removed].copy()
                result['mapping_status'] = 'matched'
                result['match_type'] = 'suffix_normalized'
                return result

            # Try normalized match on suffix-removed string
            normalized_suffix_removed = self._normalize(suffix_removed)
            if normalized_suffix_removed in self.mapping_dict:
                result = self.mapping_dict[normalized_suffix_removed].copy()
                result['mapping_status'] = 'matched'
                result['match_type'] = 'suffix_normalized'
                return result

        # 4. Unmatched
        return {
            'coverage_code': '',
            'coverage_name_canonical': '',
            'mapping_status': 'unmatched',
            'match_type': 'none'
        }


def map_scope_to_canonical(
    scope_csv_path: str,
    mapping_excel_path: str,
    output_csv_path: str
) -> Dict:
    """
    Scope CSV를 canonical mapping하여 저장

    Args:
        scope_csv_path: 입력 scope CSV 경로
        mapping_excel_path: 담보명 mapping 엑셀 경로
        output_csv_path: 출력 mapped CSV 경로

    Returns:
        dict: 매핑 통계
    """
    mapper = CanonicalMapper(mapping_excel_path)

    # Scope CSV 읽기
    scope_rows = []
    with open(scope_csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        scope_rows = list(reader)

    # 매핑 실행
    mapped_rows = []
    stats = {'matched': 0, 'unmatched': 0}

    for row in scope_rows:
        coverage_name_raw = row['coverage_name_raw']
        mapping_result = mapper.map_coverage(coverage_name_raw)

        # STEP NEXT-34-ε: Preserve search_key from Step1 (if exists)
        mapped_row = {
            'coverage_name_raw': coverage_name_raw,
            'insurer': row['insurer'],
            'source_page': row['source_page'],
            'coverage_code': mapping_result['coverage_code'],
            'coverage_name_canonical': mapping_result['coverage_name_canonical'],
            'mapping_status': mapping_result['mapping_status'],
            'match_type': mapping_result.get('match_type', '')
        }
        # Preserve search_key if present in Step1 output
        if 'coverage_name_search_key' in row:
            mapped_row['coverage_name_search_key'] = row['coverage_name_search_key']

        mapped_rows.append(mapped_row)

        stats[mapping_result['mapping_status']] += 1

    # 결과 저장
    # STEP NEXT-34-ε: Include search_key in output if present
    fieldnames = [
        'coverage_name_raw', 'insurer', 'source_page',
        'coverage_code', 'coverage_name_canonical',
        'mapping_status', 'match_type'
    ]
    # Add search_key to fieldnames if it exists in any row
    if mapped_rows and 'coverage_name_search_key' in mapped_rows[0]:
        fieldnames.append('coverage_name_search_key')

    with open(output_csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(mapped_rows)

    return stats


def main():
    """CLI 실행"""
    import argparse

    parser = argparse.ArgumentParser(description='Canonical mapping')
    parser.add_argument('--insurer', type=str, default='samsung', help='보험사명')
    args = parser.parse_args()

    # 경로 설정
    base_dir = Path(__file__).parent.parent.parent
    insurer = args.insurer

    scope_csv = base_dir / "data" / "scope" / f"{insurer}_scope.csv"
    mapping_excel = base_dir / "data" / "sources" / "mapping" / "담보명mapping자료.xlsx"
    output_csv = base_dir / "data" / "scope" / f"{insurer}_scope_mapped.csv"

    print(f"[Step 2] Canonical Mapping")
    print(f"[Step 2] Input: {scope_csv}")
    print(f"[Step 2] Mapping source: {mapping_excel}")
    print(f"[Step 2] Output: {output_csv}")

    # 매핑 실행
    stats = map_scope_to_canonical(
        str(scope_csv),
        str(mapping_excel),
        str(output_csv)
    )

    print(f"\n[Step 2] Mapping completed:")
    print(f"  - Matched: {stats['matched']}")
    print(f"  - Unmatched: {stats['unmatched']}")
    print(f"  - Total: {stats['matched'] + stats['unmatched']}")
    print(f"\n✓ Output: {output_csv}")


if __name__ == "__main__":
    main()
